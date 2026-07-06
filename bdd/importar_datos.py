import sqlite3
import re
import os

DB = os.path.join(os.path.dirname(__file__), 'expedientes.db')
SCHEMA = os.path.join(os.path.dirname(__file__), 'Tablas8.sql')
DATA = os.path.join(os.path.dirname(__file__), 'datos_excel.sql')
EXCEL = os.path.join(os.path.dirname(__file__), 'CONTROL DE DOCUMENTOS JUNIO 2026.xlsx')

COLUMNS = [
    'solped', 'id_gerencia', 'id_superintendencia', 'id_emisor', 'id_documento',
    'fecha_presupuesto_base', 'presupuesto_base_usd', 'tipo_cambio', 'presupuesto_base_bs',
    'id_plan', 'descripcion_proceso', 'id_modalidad', 'id_art', 'id_tipo_contrato',
    'nro_acta_apertura', 'cantidad_frentes', 'nro_resolucion_jd', 'id_estatus',
    'fecha_recibido', 'fecha_devuelto', 'id_receptor', 'nro_proceso', 'id_resultado',
    'nro_contrato_sicac', 'nro_contrato_sap', 'id_empresa', 'tiempo_ejecucion',
    'monto_adjudicado_bs', 'monto_adjudicado_usd', 'fecha_firma_contrato', 'observaciones_generales'
]

def parse_row_values(line):
    line = line.strip()
    if line.endswith(';'):
        line = line[:-1]
    if line.endswith(','):
        line = line[:-1]
    if line.startswith('(') and line.endswith(')'):
        line = line[1:-1]
    else:
        return None

    vals = []
    current = ''
    in_quote = False
    depth = 0
    for ch in line:
        if ch == "'" and not in_quote:
            in_quote = True
            current += ch
        elif ch == "'" and in_quote:
            in_quote = False
            current += ch
        elif ch == '(' and not in_quote:
            depth += 1
            current += ch
        elif ch == ')' and not in_quote:
            depth -= 1
            current += ch
        elif ch == ',' and depth == 0 and not in_quote:
            vals.append(current.strip())
            current = ''
        else:
            current += ch
    if current.strip():
        vals.append(current.strip())
    return vals

def parse_value(v):
    if v.upper() == 'NULL':
        return None
    if v.startswith("'") and v.endswith("'"):
        return v[1:-1]
    if v == '':
        return None
    try:
        if '.' in v:
            return float(v)
        return int(v)
    except ValueError:
        return v

def main():
    if os.path.exists(DB):
        os.remove(DB)

    con = sqlite3.connect(DB)
    con.execute("PRAGMA foreign_keys = OFF")
    cur = con.cursor()

    with open(SCHEMA) as f:
        cur.executescript(f.read())

    with open(DATA) as f:
        data_sql = f.read()

    catalog_sections = data_sql.split('INSERT INTO expedientes')[0]
    cur.executescript(catalog_sections)

    rows_part = 'INSERT INTO expedientes' + data_sql.split('INSERT INTO expedientes')[1]
    lines = rows_part.split('\n')

    all_rows = []
    for line in lines:
        raw = parse_row_values(line)
        if raw is not None and len(raw) == len(COLUMNS):
            row = [parse_value(v) for v in raw]
            all_rows.append(row)

    # Leer observaciones del Excel (Col 20 = OBSERVACIONES HISTORIAL, Col 34 = OBSERVACIONES)
    if os.path.exists(EXCEL):
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL, data_only=True)
        ws = wb.active
        obs_col = COLUMNS.index('observaciones_generales')
        for i, row in enumerate(all_rows):
            hist = str(ws.cell(i + 2, 20).value or '').strip()
            general = str(ws.cell(i + 2, 34).value or '').strip()
            partes = [p for p in [hist, general] if p]
            if partes:
                row[obs_col] = '\n'.join(partes)

    insert_cols = ', '.join(COLUMNS)
    placeholders = ', '.join(['?' for _ in COLUMNS])
    update_set = ', '.join([f"{c} = ?" for c in COLUMNS])

    seen_empty = False
    insert_count = 0
    update_count = 0
    solped_to_id = {}

    for vals in all_rows:
        solped = vals[0]

        if solped and solped.strip():
            if solped in solped_to_id:
                existing_id = solped_to_id[solped]
                vals.append(existing_id)
                try:
                    cur.execute(f"UPDATE expedientes SET {update_set} WHERE id_expediente = ?", vals)
                except Exception as e:
                    print(f"UPDATE err solped={solped}: {e}")
                    print(f"  vals={vals}")
                    raise
                update_count += 1
            else:
                try:
                    cur.execute(f"INSERT INTO expedientes ({insert_cols}) VALUES ({placeholders})", vals)
                except Exception as e:
                    print(f"INSERT err solped={solped}: {e}")
                    print(f"  vals={vals}")
                    raise
                new_id = cur.lastrowid
                solped_to_id[solped] = new_id
                insert_count += 1
        else:
            if not seen_empty:
                try:
                    cur.execute(f"INSERT INTO expedientes ({insert_cols}) VALUES ({placeholders})", vals)
                except Exception as e:
                    print(f"INSERT err solped=empty: {e}")
                    print(f"  vals={vals}")
                    raise
                seen_empty = True
                insert_count += 1
            else:
                cur.execute("SELECT id_expediente FROM expedientes WHERE (solped IS NULL OR solped = '') LIMIT 1")
                existing = cur.fetchone()
                if existing:
                    existing_id = existing[0]
                    vals.append(existing_id)
                    try:
                        cur.execute(f"UPDATE expedientes SET {update_set} WHERE id_expediente = ?", vals)
                    except Exception as e:
                        print(f"UPDATE err solped=empty: {e}")
                        print(f"  vals={vals}")
                        raise
                    update_count += 1

    con.commit()

    hist_count = cur.execute("SELECT COUNT(*) FROM historial_movimientos").fetchone()[0]
    exp_count = cur.execute("SELECT COUNT(*) FROM expedientes").fetchone()[0]

    print(f"Expedientes: {exp_count} (insert={insert_count}, update={update_count})")
    print(f"Historial movimientos: {hist_count}")

    dups = cur.execute("SELECT solped, COUNT(*) FROM expedientes GROUP BY solped HAVING COUNT(*) > 1").fetchall()
    if dups:
        print(f"\nATENCIÓN: {len(dups)} solpeds con duplicados:")
        for s, c in dups:
            print(f"  {s}: {c} veces")

    print("\n--- Muestra de observaciones ---")
    for row in cur.execute("SELECT id_expediente, solped, observaciones_generales FROM expedientes WHERE observaciones_generales IS NOT NULL LIMIT 3"):
        print(f"  id={row[0]} solped={row[1]}: {row[2][:80]}...")
    for row in cur.execute("SELECT id_expediente, solped, observaciones_generales FROM historial_movimientos WHERE observaciones_generales IS NOT NULL AND observaciones_generales != '' LIMIT 3"):
        print(f"  HIST id={row[0]} solped={row[1]}: {row[2][:80]}...")

    con.close()

if __name__ == '__main__':
    main()
