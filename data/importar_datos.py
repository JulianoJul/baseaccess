import sqlite3
import re
import os

DB = os.path.join(os.path.dirname(__file__), 'expedientes.db')
SCHEMA_FILES = [
    os.path.join(os.path.dirname(__file__), 'sql', '01_master_control_docs_presidencia.sql'),
    os.path.join(os.path.dirname(__file__), 'sql', '02_modulos_adicionales.sql'),
    os.path.join(os.path.dirname(__file__), 'sql', '03_ruta_procesos.sql'),
]
DATA_FILES = [
    os.path.join(os.path.dirname(__file__), 'datos_excel.sql'),
    os.path.join(os.path.dirname(__file__), 'Datos_excel2.sql'),
]
EXCEL = os.path.join(os.path.dirname(__file__), 'CONTROL DE DOCUMENTOS JUNIO 2026.xlsx')

COLUMNS = [
    'solped', 'id_gerencia', 'id_superintendencia', 'id_emisor', 'id_documento',
    'fecha_presupuesto_base', 'presupuesto_base_usd', 'tipo_cambio', 'presupuesto_base_bs',
    'id_plan', 'descripcion_proceso', 'id_modalidad', 'id_art', 'id_tipo_contrato',
    'nro_acta_apertura', 'cantidad_frentes', 'nro_resolucion_jd', 'id_estatus',
    'fecha_recibido', 'fecha_devuelto', 'id_receptor', 'nro_proceso', 'id_resultado',
    'nro_contrato_sicac', 'nro_contrato_sap', 'id_empresa', 'tiempo_ejecucion',
    'monto_adjudicado_bs', 'monto_adjudicado_usd', 'fecha_firma_contrato', 'observaciones', 'notas',
    'fecha_creacion', 'fecha_actualizacion'
]

def parse_row_values(line):
    line = line.strip()
    if line.endswith(';'):
        line = line[:-1]
    if line.endswith(','):
        line = line[:-1]
    if line.startswith('(') and line.endswith(')'):
        line = line[1:-1]
    else:
        return None

    vals = []
    current = ''
    in_quote = False
    depth = 0
    for ch in line:
        if ch == "'" and not in_quote:
            in_quote = True
            current += ch
        elif ch == "'" and in_quote:
            in_quote = False
            current += ch
        elif ch == '(' and not in_quote:
            depth += 1
            current += ch
        elif ch == ')' and not in_quote:
            depth -= 1
            current += ch
        elif ch == ',' and depth == 0 and not in_quote:
            vals.append(current.strip())
            current = ''
        else:
            current += ch
    if current.strip():
        vals.append(current.strip())
    return vals

def parse_value(v):
    if v.upper() == 'NULL':
        return None
    if v.startswith("'") and v.endswith("'"):
        return v[1:-1]
    if v == '':
        return None
    try:
        if '.' in v:
            return float(v)
        return int(v)
    except ValueError:
        return v

def main():
    if os.path.exists(DB):
        os.remove(DB)

    con = sqlite3.connect(DB)
    con.execute("PRAGMA foreign_keys = OFF")
    cur = con.cursor()

    for schema_file in SCHEMA_FILES:
        with open(schema_file) as f:
            cur.executescript(f.read())

    # Procesar datos_excel.sql (catálogos + expedientes)
    with open(DATA_FILES[0]) as f:
        data_sql = f.read()

    catalog_sections = data_sql.split('INSERT INTO expedientes')[0]
    cur.executescript(catalog_sections)

    rows_part = 'INSERT INTO expedientes' + data_sql.split('INSERT INTO expedientes')[1]
    lines = rows_part.split('\n')

    # Ejecutar Datos_excel2.sql (otros módulos: req_materiales, memorandums, recobros, etc.)
    with open(DATA_FILES[1]) as f:
        cur.executescript(f.read())

    all_rows = []
    # Valores esperados por fila del SQL: 31 (sin notas, fecha_creacion, fecha_actualizacion)
    SQL_VALUES_COUNT = len(COLUMNS) - 3
    for line in lines:
        raw = parse_row_values(line)
        if raw is not None and len(raw) == SQL_VALUES_COUNT:
            row = [parse_value(v) for v in raw]
            row.append(None)   # notas
            row.append(None)   # fecha_creacion (se setea abajo desde Excel)
            row.append(None)   # fecha_actualizacion (se setea abajo desde Excel)
            all_rows.append(row)

    # Leer observaciones y notas del Excel (Col 20 = OBSERVACIONES HISTORIAL, Col 34 = OBSERVACIONES)
    # Trackear por solped:
    #   fecha_creacion = MIN(fecha_recibido)  — cuando nace el expediente
    #   fecha_actualizacion = MAX(fecha_devuelto, fecha_recibido)  — último movimiento
    solped_fechas = {}  # solped -> [min_recibido, max_actualizacion]
    if os.path.exists(EXCEL):
        import openpyxl
        import datetime
        wb = openpyxl.load_workbook(EXCEL, data_only=True)
        ws = wb.active
        obs_col = COLUMNS.index('observaciones')
        notas_col = COLUMNS.index('notas')

        def to_date_str(val):
            if isinstance(val, (datetime.datetime, datetime.date)):
                return val.strftime('%Y-%m-%d')
            s = str(val).strip() if val else ''
            import re
            m = re.match(r'(\d{2})/(\d{2})/(\d{4})', s)
            if m:
                d, mo, y = m.groups()
                return f'{y}-{mo}-{d}'
            return None

        for i, row in enumerate(all_rows):
            hist = str(ws.cell(i + 2, 20).value or '').strip()
            if hist:
                row[obs_col] = hist
            notas = str(ws.cell(i + 2, 34).value or '').strip()
            if notas:
                row[notas_col] = notas
            fecha_rec = to_date_str(ws.cell(i + 2, 22).value)
            fecha_dev = to_date_str(ws.cell(i + 2, 23).value)
            # fecha_actualizacion candidata = la más nueva entre recibido y devuelto de esta fila
            cand_actualizacion = fecha_dev or fecha_rec
            solped_key = (row[0] or '').strip() if row[0] else ''
            prev = solped_fechas.get(solped_key)
            if prev is None:
                solped_fechas[solped_key] = [fecha_rec, cand_actualizacion]
            else:
                if fecha_rec and (prev[0] is None or fecha_rec < prev[0]):
                    prev[0] = fecha_rec
                if cand_actualizacion and (prev[1] is None or cand_actualizacion > prev[1]):
                    prev[1] = cand_actualizacion

    insert_cols = ', '.join(COLUMNS)
    placeholders = ', '.join(['?' for _ in COLUMNS])
    update_set = ', '.join([f"{c} = ?" for c in COLUMNS])

    seen_empty = False
    insert_count = 0
    update_count = 0
    solped_to_id = {}

    for vals in all_rows:
        solped = vals[0]

        if solped and solped.strip():
            if solped in solped_to_id:
                existing_id = solped_to_id[solped]
                existing_obs = cur.execute("SELECT observaciones FROM expedientes WHERE id_expediente = ?", (existing_id,)).fetchone()[0]
                if vals[30] and existing_obs:
                    vals[30] = existing_obs + '\n' + vals[30]
                vals.append(existing_id)
                try:
                    cur.execute(f"UPDATE expedientes SET {update_set} WHERE id_expediente = ?", vals)
                except Exception as e:
                    print(f"UPDATE err solped={solped}: {e}")
                    print(f"  vals={vals}")
                    raise
                update_count += 1
            else:
                try:
                    cur.execute(f"INSERT INTO expedientes ({insert_cols}) VALUES ({placeholders})", vals)
                except Exception as e:
                    print(f"INSERT err solped={solped}: {e}")
                    print(f"  vals={vals}")
                    raise
                new_id = cur.lastrowid
                solped_to_id[solped] = new_id
                insert_count += 1
        else:
            if not seen_empty:
                try:
                    cur.execute(f"INSERT INTO expedientes ({insert_cols}) VALUES ({placeholders})", vals)
                except Exception as e:
                    print(f"INSERT err solped=empty: {e}")
                    print(f"  vals={vals}")
                    raise
                seen_empty = True
                insert_count += 1
            else:
                cur.execute("SELECT id_expediente FROM expedientes WHERE (solped IS NULL OR solped = '') LIMIT 1")
                existing = cur.fetchone()
                if existing:
                    existing_id = existing[0]
                    existing_obs = cur.execute("SELECT observaciones FROM expedientes WHERE id_expediente = ?", (existing_id,)).fetchone()[0]
                    if vals[30] and existing_obs:
                        vals[30] = existing_obs + '\n' + vals[30]
                    vals.append(existing_id)
                    try:
                        cur.execute(f"UPDATE expedientes SET {update_set} WHERE id_expediente = ?", vals)
                    except Exception as e:
                        print(f"UPDATE err solped=empty: {e}")
                        print(f"  vals={vals}")
                        raise
                    update_count += 1

    con.commit()

    # Setear fecha_creacion y fecha_actualizacion según el tracking por solped:
    #   fecha_creacion = MIN(fecha_recibido) — nacimiento del expediente
    #   fecha_actualizacion = MAX(fecha_devuelto, fecha_recibido) — último movimiento
    for solped_key, (min_rec, max_act) in solped_fechas.items():
        if not min_rec and not max_act:
            continue
        if solped_key:
            cur.execute(
                "UPDATE expedientes SET fecha_creacion = ?, fecha_actualizacion = ? WHERE solped = ?",
                (min_rec, max_act, solped_key)
            )
        else:
            cur.execute(
                "UPDATE expedientes SET fecha_creacion = ?, fecha_actualizacion = ? WHERE (solped IS NULL OR solped = '')",
                (min_rec, max_act)
            )
    upd_count = cur.rowcount
    con.commit()
    print(f"fecha_actualizacion recalculada para {upd_count} expedientes (último movimiento del historial)")

    hist_count = cur.execute("SELECT COUNT(*) FROM historial_movimientos").fetchone()[0]
    exp_count = cur.execute("SELECT COUNT(*) FROM expedientes").fetchone()[0]

    print(f"Expedientes: {exp_count} (insert={insert_count}, update={update_count})")
    print(f"Historial movimientos: {hist_count}")

    dups = cur.execute("SELECT solped, COUNT(*) FROM expedientes GROUP BY solped HAVING COUNT(*) > 1").fetchall()
    if dups:
        print(f"\nATENCIÓN: {len(dups)} solpeds con duplicados:")
        for s, c in dups:
            print(f"  {s}: {c} veces")

    print("\n--- Muestra de observaciones ---")
    for row in cur.execute("SELECT id_expediente, solped, observaciones FROM expedientes WHERE observaciones IS NOT NULL LIMIT 3"):
        print(f"  id={row[0]} solped={row[1]}: {row[2][:80]}...")
    for row in cur.execute("SELECT id_expediente, solped, observaciones FROM historial_movimientos WHERE observaciones IS NOT NULL AND observaciones != '' LIMIT 3"):
        print(f"  HIST id={row[0]} solped={row[1]}: {row[2][:80]}...")

    con.close()

if __name__ == '__main__':
    main()
